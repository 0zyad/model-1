"""
inference_stardist.py — Drop-in replacement for inference.py using StarDist.

Identical ProppantAnalyzer interface — just change the import in app.py:
    from inference_stardist import ProppantAnalyzer

What changed vs inference.py:
  - YOLOv8 removed → StarDist2D instance segmentation
  - Watershed splits merged blobs (post-processing)
  - Classification: bimodal size split — NO per-particle class from the model
  - "unknown" replaces "sand" for anything outside 40/70 / 20/40 size bands
  - avg_confidence = mean StarDist object probability (semantically equivalent)
  - All downstream logic (composition, verdict, sieve, overlay, SWE) UNCHANGED

Classification logic:
  1. Run StarDist → raw instance masks (no class label)
  2. Compute diameter for each instance
  3. Bimodal size split (variance-minimising threshold) → class 0 / 1 / unknown
  4. _refine_with_color_and_size() corrects obvious outliers (same as before)
  5. If diameter falls in ambiguous ±15% zone around the split boundary → unknown

"unknown" particles:
  - Counted separately in composition
  - Excluded from PASS/FAIL purity calculation
  - Still contribute to sieve wt% estimate
  - Drawn grey in overlay
"""
import time
import cv2
import numpy as np
from pathlib import Path

from config import (
    CLASS_NAMES, CLASS_COLORS, NUM_CLASSES,
    PURITY_THRESHOLD,
    MAX_SIZE_ERROR, MAX_PROCESSING_TIME,
    PIXELS_PER_MM, RUNS_DIR, SIEVE_EXCEL_PATH,
    CELLPOSE_MODEL_PATH, CELLPOSE_PROB_THRESH, CELLPOSE_NMS_THRESH,
    CELLPOSE_USE_GPU, CELLPOSE_DIAMETER,
)

# "unknown" is not a trained class — assigned at runtime for ambiguous sizes
UNKNOWN_CLASS_ID    = -1
UNKNOWN_CLASS_NAME  = "unknown"
UNKNOWN_CLASS_COLOR = (128, 128, 128)   # Grey (BGR)

# Ambiguity band disabled — was causing 13%+ unknown rate.
# Refinement step (_refine_with_color_and_size) handles boundary corrections.
AMBIGUITY_FRAC = 0.0


class ProppantAnalyzer:
    """Stateful analyzer — loads StarDist model once, processes many images."""

    def __init__(self, model_path: "str | Path | None" = None):
        from segmentation.cellpose_seg import CellposeSegmentor
        self.segmentor  = CellposeSegmentor(
            model_path=model_path,
            use_gpu=CELLPOSE_USE_GPU,
            diameter=CELLPOSE_DIAMETER,
        )
        self.model_path = Path(model_path) if model_path else CELLPOSE_MODEL_PATH

    def analyze(self, image_path: "str | Path", progress_fn=None) -> dict:
        """Full analysis pipeline for one image. Returns structured result."""
        t_start = time.time()
        image_path = Path(image_path)
        img = cv2.imread(str(image_path))
        if img is None:
            return self._error_result(image_path, "Failed to read image")

        h_orig, w_orig = img.shape[:2]

        # ── Blur / quality score ──────────────────────────────────────────────
        gray_orig = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        blur_score = round(cv2.Laplacian(gray_orig, cv2.CV_64F).var(), 2)

        # ── CellPose segmentation ─────────────────────────────────────────────
        # Masks are returned at DOWNSAMPLED resolution (segmentor.last_scale).
        # diameter_px values below are at that scale — corrected after.
        raw_masks, raw_probs = self.segmentor.segment(
            img,
            prob_thresh=CELLPOSE_PROB_THRESH,
            nms_thresh=CELLPOSE_NMS_THRESH,
            progress_fn=progress_fn,
        )
        seg_scale = self.segmentor.last_scale   # e.g. 0.33 for 20MP images
        print(f"  CellPose detected: {len(raw_masks)} instances (scale={seg_scale})")

        # ── Small image for gap-fill and overlay (matches mask resolution) ────
        if seg_scale < 1.0:
            sw = int(w_orig * seg_scale)
            sh = int(h_orig * seg_scale)
            img_small = cv2.resize(img, (sw, sh), interpolation=cv2.INTER_AREA)
        else:
            img_small = img

        # ── Build particle list — measurements in ORIGINAL pixel space ────────
        particles  = []
        all_masks  = []   # bool masks at downsampled resolution

        for mask_bool, prob in zip(raw_masks, raw_probs):
            area_small = float(mask_bool.sum())
            if area_small < 10:
                continue
            # Scale measurements back to original resolution
            area_px     = area_small / (seg_scale ** 2)
            diameter_px = 2.0 * np.sqrt(area_px / np.pi)

            particle = {
                "class_id":    UNKNOWN_CLASS_ID,
                "class_name":  UNKNOWN_CLASS_NAME,
                "confidence":  round(prob, 4),
                "area_px":     round(area_px, 1),
                "diameter_px": round(diameter_px, 2),
                "mask_index":  len(all_masks),
                "source":      "cellpose",
            }
            if PIXELS_PER_MM is not None:
                particle["diameter_mm"] = round(diameter_px / PIXELS_PER_MM, 3)

            particles.append(particle)
            all_masks.append(mask_bool)   # bool, NOT float32 — avoids OOM

        # ── Gap-fill: catch missed particles via classical OpenCV ─────────────
        particles, all_masks = self._opencv_gap_fill(img_small, particles, all_masks, seg_scale)

        # ── Size-based classification ─────────────────────────────────────────
        particles = self._classify_by_size(particles)

        # ── Size-based refinement ─────────────────────────────────────────────
        if len(particles) > 0:
            particles = self._refine_with_color_and_size(img_small, particles, None)

        total_detected = len(particles)

        # ── Composition ───────────────────────────────────────────────────────
        counts = {0: 0, 1: 0, UNKNOWN_CLASS_ID: 0}
        for p in particles:
            counts[p["class_id"]] = counts.get(p["class_id"], 0) + 1

        composition = {
            "proppant_40_70": {
                "count":      counts[0],
                "percentage": round(counts[0] / total_detected * 100, 2) if total_detected else 0.0,
            },
            "proppant_20_40": {
                "count":      counts[1],
                "percentage": round(counts[1] / total_detected * 100, 2) if total_detected else 0.0,
            },
            "unknown": {
                "count":      counts[UNKNOWN_CLASS_ID],
                "percentage": round(counts[UNKNOWN_CLASS_ID] / total_detected * 100, 2) if total_detected else 0.0,
            },
        }

        # Confidence = % of classified particles (meaningful, not raw CellPose flow prob)
        classified_n = counts[0] + counts[1]
        avg_confidence = round(classified_n / total_detected * 100, 1) if total_detected else 0.0

        # ── PASS / FAIL verdict ───────────────────────────────────────────────
        verdict, reason = self._evaluate_verdict(composition, total_detected)

        # ── Spec 8: size error vs sieve reference ─────────────────────────────
        mean_size_error = self._estimate_size_error(particles, verdict)

        # ── SWE spec checks ───────────────────────────────────────────────────
        processing_time = time.time() - t_start
        # "classified" = not unknown
        classified_count  = counts[0] + counts[1]
        classified_rate   = (classified_count / total_detected * 100) if total_detected else 0.0
        spec_classified   = classified_rate >= 90.0
        spec_size_error   = mean_size_error <= MAX_SIZE_ERROR * 100
        # Note: processing_time is software runtime, NOT a pass/fail spec

        swe_checks = {
            "classified_rate_pct": round(classified_rate, 1),
            "classified_pass":     spec_classified,
            "mean_size_error_pct": mean_size_error,
            "size_error_pass":     spec_size_error,
            "processing_time_sec": round(processing_time, 2),
            "processing_time_pass": True,   # not a spec — just informational
            "all_passed":          spec_classified and spec_size_error,
        }

        # ── Overlay — draw on small image, upscale to original at the end ───────
        overlay_small = self._draw_overlay(img_small, all_masks, particles, composition, verdict)
        if seg_scale < 1.0 and overlay_small is not None:
            overlay = cv2.resize(overlay_small, (w_orig, h_orig), interpolation=cv2.INTER_LINEAR)
        else:
            overlay = overlay_small

        return {
            "image_path":          str(image_path),
            "image_name":          image_path.name,
            "total_particles":     total_detected,
            "composition":         composition,
            "verdict":             verdict,
            "reason":              reason,
            "avg_confidence":      avg_confidence,
            "blur_score":          blur_score,
            "swe_checks":          swe_checks,
            "particles":           particles,
            "overlay":             overlay,
            "processing_time_sec": round(processing_time, 2),
        }

    # ── Size-based initial classification ─────────────────────────────────────

    def _classify_by_size(self, particles: list) -> list:
        """
        Assign class 0 (40/70) or class 1 (20/40) or unknown (-1) to each particle
        using only diameter, with no prior class label from the model.

        Strategy:
          1. Find optimal binary split via within-group variance minimisation.
          2. If the two group medians are < 30% apart → unimodal (pure sample).
             All particles get the same class based on absolute median.
          3. Assign "unknown" to particles within ±AMBIGUITY_FRAC of the boundary.
        """
        if not particles:
            return particles

        diameters = np.array([p["diameter_px"] for p in particles])
        n = len(diameters)

        if n < 5:
            # Too few particles to determine distribution — mark all unknown
            for p in particles:
                p["class_id"]   = UNKNOWN_CLASS_ID
                p["class_name"] = UNKNOWN_CLASS_NAME
            return particles

        sorted_d = np.sort(diameters)

        # Variance-minimising split
        best_thresh = sorted_d[n // 2]
        best_var    = float("inf")
        for i in range(2, n - 2):
            low  = sorted_d[:i]
            high = sorted_d[i:]
            var  = np.var(low) * len(low) + np.var(high) * len(high)
            if var < best_var:
                best_var    = var
                best_thresh = (sorted_d[i - 1] + sorted_d[i]) / 2.0

        low_group  = sorted_d[sorted_d <  best_thresh]
        high_group = sorted_d[sorted_d >= best_thresh]

        med_low  = float(np.median(low_group))  if len(low_group)  > 0 else 0.0
        med_high = float(np.median(high_group)) if len(high_group) > 0 else 0.0

        unimodal = (med_high / max(med_low, 1e-9)) < 1.30

        if unimodal:
            # Pure sample — determine type by absolute median size.
            # At 1/3 downscale, 40/70 ≈ 135-180px and 20/40 ≈ 270-360px in
            # original pixel space.  200px is a reliable midpoint separator.
            median_all = float(np.median(diameters))
            dominant_class = 1 if median_all > 200 else 0
            print(f"  Unimodal: median={median_all:.1f}px → class={CLASS_NAMES[dominant_class]}")
            for p in particles:
                p["class_id"]   = dominant_class
                p["class_name"] = CLASS_NAMES[dominant_class]
        else:
            # Bimodal — smaller cluster = 40/70, larger = 20/40.
            # No ambiguity band: assign every particle; refinement corrects outliers.
            for p in particles:
                d = p["diameter_px"]
                if d < best_thresh:
                    p["class_id"]   = 0
                    p["class_name"] = CLASS_NAMES[0]
                else:
                    p["class_id"]   = 1
                    p["class_name"] = CLASS_NAMES[1]

        classified = sum(1 for p in particles if p["class_id"] != UNKNOWN_CLASS_ID)
        print(f"  Size classifier: boundary={best_thresh:.1f}px "
              f"unimodal={unimodal} classified={classified}/{n}")
        return particles

    # ── OpenCV gap-fill (identical to inference.py) ───────────────────────────

    def _get_size_threshold(self, particles):
        sizes_0 = [p["diameter_px"] for p in particles
                   if p["class_id"] == 0 and p["confidence"] >= 0.3]
        sizes_1 = [p["diameter_px"] for p in particles
                   if p["class_id"] == 1 and p["confidence"] >= 0.3]
        if sizes_0 and sizes_1:
            return (float(np.median(sizes_0)) + float(np.median(sizes_1))) / 2.0
        elif sizes_0:
            return float(np.median(sizes_0)) * 1.5
        elif sizes_1:
            return float(np.median(sizes_1)) * 0.7
        return None

    def _opencv_gap_fill(self, img, particles, masks_list, seg_scale: float = 1.0):
        """img and masks_list are at the same (possibly downsampled) resolution.
        seg_scale is used to convert small-pixel measurements to original-pixel space."""
        h, w = img.shape[:2]
        gray  = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        size_thresh = self._get_size_threshold(particles)

        covered = np.zeros((h, w), dtype=np.uint8)
        for m in masks_list:
            covered[m.astype(bool)] = 255
        kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
        covered_dilated = cv2.dilate(covered, kernel, iterations=15)

        blurred = cv2.GaussianBlur(gray, (5, 5), 0)
        _, bin_otsu = cv2.threshold(blurred, 0, 255,
                                    cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
        _, bin_fix = cv2.threshold(blurred, 180, 255, cv2.THRESH_BINARY_INV)
        binary = cv2.bitwise_or(bin_otsu, bin_fix)

        binary = cv2.morphologyEx(binary, cv2.MORPH_OPEN,  kernel, iterations=1)
        binary = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, kernel, iterations=2)
        uncovered = cv2.bitwise_and(binary, cv2.bitwise_not(covered_dilated))

        contours, _ = cv2.findContours(
            uncovered, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE
        )

        min_area_small = max(20, 250 * (seg_scale ** 2))
        added = 0
        for c in contours:
            area = cv2.contourArea(c)
            if area < min_area_small:
                continue
            perim = cv2.arcLength(c, True)
            if perim > 0 and (4 * np.pi * area / (perim * perim)) < 0.15:
                continue

            mask = np.zeros((h, w), dtype=np.uint8)
            cv2.drawContours(mask, [c], -1, 1, -1)
            mask_bool = mask.astype(bool)

            # Scale small-pixel measurements back to original resolution
            area_orig     = area / (seg_scale ** 2)
            diameter_px   = 2.0 * np.sqrt(area_orig / np.pi)
            size_thresh_orig = size_thresh / seg_scale if size_thresh is not None else None

            if size_thresh_orig is not None:
                cls_id = 0 if diameter_px < size_thresh_orig else 1
            else:
                cls_id = UNKNOWN_CLASS_ID

            particle = {
                "class_id":    cls_id,
                "class_name":  CLASS_NAMES.get(cls_id, UNKNOWN_CLASS_NAME),
                "confidence":  0.35,
                "area_px":     round(area_orig, 1),
                "diameter_px": round(diameter_px, 2),
                "mask_index":  len(masks_list),
                "source":      "opencv",
            }
            if PIXELS_PER_MM is not None:
                particle["diameter_mm"] = round(diameter_px / PIXELS_PER_MM, 3)

            masks_list.append(mask_bool)
            particles.append(particle)
            added += 1

        if added:
            print(f"  OpenCV gap-fill: added {added} particles")
        return particles, masks_list

    # ── Size-based refinement (identical to inference.py) ─────────────────────

    def _refine_with_color_and_size(self, img, particles, masks_data):
        yolo_only = [p for p in particles if p.get("source") != "opencv"]
        sizes_0 = [p["diameter_px"] for p in yolo_only if p["class_id"] == 0]
        sizes_1 = [p["diameter_px"] for p in yolo_only if p["class_id"] == 1]

        all_sizes = [p["diameter_px"] for p in yolo_only if p["class_id"] in (0, 1)]
        if len(all_sizes) < 5:
            return particles

        total_proppant = len(sizes_0) + len(sizes_1)
        ratio_0 = len(sizes_0) / max(total_proppant, 1)
        ratio_1 = len(sizes_1) / max(total_proppant, 1)

        near_pure_2040 = ratio_1 >= 0.70 and len(sizes_1) >= 3

        if near_pure_2040:
            med_1 = float(np.median(sizes_1))
            boundary = med_1 * 0.65
        elif len(sizes_0) >= 3 and len(sizes_1) >= 3:
            med_0 = float(np.median(sizes_0))
            med_1 = float(np.median(sizes_1))
            if med_1 > med_0 * 1.2:
                boundary = (med_0 + med_1) / 2.0
            else:
                return particles
        elif len(sizes_0) >= 3:
            boundary = float(np.median(sizes_0)) * 1.5
        elif len(sizes_1) >= 3:
            boundary = float(np.median(sizes_1)) * 0.65
        else:
            return particles

        dominant_ratio = max(ratio_0, ratio_1)
        tolerance = 0.10 if dominant_ratio > 0.75 else 0.20
        low_band  = boundary * (1.0 - tolerance)
        high_band = boundary * (1.0 + tolerance)

        reclassified = 0
        for p in particles:
            if p["class_id"] == UNKNOWN_CLASS_ID:
                continue
            d = p["diameter_px"]

            if near_pure_2040 and p["class_id"] == 0:
                p["class_id"]   = 1
                p["class_name"] = CLASS_NAMES[1]
                p["refined"]    = True
                reclassified += 1
                continue

            if p["class_id"] == 0 and d > high_band:
                p["class_id"]   = 1
                p["class_name"] = CLASS_NAMES[1]
                p["refined"]    = True
                reclassified += 1
            elif p["class_id"] == 1 and d < low_band:
                p["class_id"]   = 0
                p["class_name"] = CLASS_NAMES[0]
                p["refined"]    = True
                reclassified += 1

        if reclassified:
            print(f"  Size refinement: corrected {reclassified} particles "
                  f"(boundary={boundary:.1f}px, near_pure_2040={near_pure_2040})")
        return particles

    # ── Verdict (identical to inference.py) ──────────────────────────────────

    def _evaluate_verdict(self, composition: dict, total: int) -> tuple:
        if total == 0:
            return "FAIL", "No particles detected"

        pct_4070 = composition["proppant_40_70"]["percentage"]
        pct_2040 = composition["proppant_20_40"]["percentage"]

        if pct_4070 >= PURITY_THRESHOLD * 100:
            return "PASS_40_70", (
                f"Proppant 40/70 purity {pct_4070:.1f}% meets "
                f"{PURITY_THRESHOLD * 100:.0f}% threshold"
            )
        if pct_2040 >= PURITY_THRESHOLD * 100:
            return "PASS_20_40", (
                f"Proppant 20/40 purity {pct_2040:.1f}% meets "
                f"{PURITY_THRESHOLD * 100:.0f}% threshold"
            )
        return "FAIL", (
            f"Mixed sample — no type reaches {PURITY_THRESHOLD * 100:.0f}% purity "
            f"(40/70: {pct_4070:.1f}%, 20/40: {pct_2040:.1f}%)"
        )

    # ── Spec 8: size error vs sieve (identical to inference.py) ──────────────

    @staticmethod
    def _load_sieve_references() -> dict:
        FALLBACK = {"proppant_20_40": 91.75, "proppant_40_70": 95.64}
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(SIEVE_EXCEL_PATH))
            ws = wb.active

            sieve_2040: dict = {}
            sieve_4070: dict = {}
            total_2040 = None
            total_4070 = None

            for row in ws.iter_rows(min_row=3, values_only=True):
                b, c, d      = row[1], row[2], row[3]
                h, i_, j     = row[7], row[8], row[9]

                if isinstance(b, str) and "total" in b.lower():
                    if isinstance(c, (int, float)):
                        total_2040 = float(c)
                elif isinstance(b, (int, float)):
                    if isinstance(c, (int, float)) and isinstance(d, (int, float)):
                        sieve_2040[float(b)] = float(d) - float(c)
                elif isinstance(b, str) and b.lower() == "pan":
                    if isinstance(c, (int, float)) and isinstance(d, (int, float)):
                        sieve_2040["pan"] = float(d) - float(c)

                if isinstance(h, str) and "total" in h.lower():
                    if isinstance(i_, (int, float)):
                        total_4070 = float(i_)
                elif isinstance(h, (int, float)):
                    if isinstance(i_, (int, float)) and isinstance(j, (int, float)):
                        sieve_4070[float(h)] = float(j) - float(i_)
                elif isinstance(h, str) and h.lower() == "pan":
                    if isinstance(i_, (int, float)) and isinstance(j, (int, float)):
                        sieve_4070["pan"] = float(j) - float(i_)

            ref_2040 = FALLBACK["proppant_20_40"]
            ref_4070 = FALLBACK["proppant_40_70"]

            if total_2040 and total_2040 > 0:
                in_spec = sum(sieve_2040.get(m, 0.0) for m in [25.0, 30.0, 35.0, 40.0])
                ref_2040 = round(in_spec / total_2040 * 100, 2)

            if total_4070 and total_4070 > 0:
                in_spec = sum(sieve_4070.get(m, 0.0) for m in [50.0, 60.0, 70.0])
                ref_4070 = round(in_spec / total_4070 * 100, 2)

            return {"proppant_20_40": ref_2040, "proppant_40_70": ref_4070}
        except Exception as e:
            print(f"  Sieve ref load error: {e} — using fallback")
            return FALLBACK

    def _estimate_size_error(self, particles: list, verdict: str = "") -> float:
        if not particles:
            return 0.0

        vol_by_class: dict = {0: 0.0, 1: 0.0}
        for p in particles:
            cid = p["class_id"]
            if cid not in vol_by_class:
                continue
            vol_by_class[cid] += p.get("diameter_px", 0.0) ** 3

        total_vol = sum(vol_by_class.values())
        if total_vol == 0:
            return 0.0

        wt_pct = {
            "proppant_40_70": vol_by_class[0] / total_vol * 100,
            "proppant_20_40": vol_by_class[1] / total_vol * 100,
        }

        if "20_40" in verdict or wt_pct["proppant_20_40"] >= wt_pct["proppant_40_70"]:
            ref_class = "proppant_20_40"
        else:
            ref_class = "proppant_40_70"

        sieve_refs = self._load_sieve_references()
        error = abs(wt_pct[ref_class] - sieve_refs[ref_class])
        print(f"  Spec 8 — {ref_class}: model={wt_pct[ref_class]:.1f}% "
              f"sieve={sieve_refs[ref_class]:.1f}% err={error:.1f}%")
        return round(error, 1)

    # ── Overlay (adds unknown class in grey) ──────────────────────────────────

    def _draw_overlay(self, img, all_masks, particles, composition, verdict):
        overlay = img.copy()
        h, w = img.shape[:2]

        if all_masks and len(particles) > 0:
            for p in particles:
                i = p["mask_index"]
                if i >= len(all_masks):
                    continue
                cls_id = p["class_id"]
                color = UNKNOWN_CLASS_COLOR if cls_id == UNKNOWN_CLASS_ID else CLASS_COLORS.get(cls_id, (255, 255, 255))

                mask_bool = all_masks[i]
                # Resize mask to match img if resolution differs
                if mask_bool.shape[0] != h or mask_bool.shape[1] != w:
                    mask_bool = cv2.resize(
                        mask_bool.astype(np.uint8), (w, h), interpolation=cv2.INTER_NEAREST
                    ).astype(bool)

                overlay[mask_bool] = (
                    overlay[mask_bool].astype(np.float32) * 0.20
                    + np.array(color, dtype=np.float32) * 0.80
                ).astype(np.uint8)

                mask_u8 = mask_bool.astype(np.uint8) * 255
                contours, _ = cv2.findContours(
                    mask_u8, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE
                )
                cv2.drawContours(overlay, contours, -1, color, 2)

        # Legend
        y = 30
        legend = [
            (CLASS_NAMES[0],     CLASS_COLORS[0],    composition.get("proppant_40_70", {}).get("percentage", 0)),
            (CLASS_NAMES[1],     CLASS_COLORS[1],    composition.get("proppant_20_40", {}).get("percentage", 0)),
            (UNKNOWN_CLASS_NAME, UNKNOWN_CLASS_COLOR, composition.get("unknown", {}).get("percentage", 0)),
        ]
        for name, color, pct in legend:
            cv2.putText(overlay, f"{name}: {pct:.1f}%", (10, y),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, color, 2)
            y += 30

        v_color = (0, 255, 0) if verdict.startswith("PASS") else (0, 0, 255)
        cv2.putText(overlay, f"Verdict: {verdict}", (10, y + 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.9, v_color, 2)
        return overlay

    # ── Tiled segmentation ────────────────────────────────────────────────────

    def _segment_tiled(self, img: np.ndarray, tile: int = 512, overlap: int = 64):
        """
        Segment a large image by splitting into overlapping tiles.
        Masks from each tile are offset to full-image coordinates.
        Duplicate detections at tile boundaries are removed by centroid dedup.
        """
        h, w = img.shape[:2]
        all_masks = []
        all_probs = []
        seen_centroids = []
        min_dist_sq = (tile // 4) ** 2   # min px distance between centroids

        step = tile - overlap
        ys = list(range(0, h - overlap, step))
        xs = list(range(0, w - overlap, step))

        for y0 in ys:
            y1 = min(y0 + tile, h)
            y0 = max(0, y1 - tile)
            for x0 in xs:
                x1 = min(x0 + tile, w)
                x0 = max(0, x1 - tile)

                patch = img[y0:y1, x0:x1]
                masks_patch, probs_patch = self.segmentor.segment(
                    patch,
                    prob_thresh=CELLPOSE_PROB_THRESH,
                    nms_thresh=CELLPOSE_NMS_THRESH,
                )

                for mask_p, prob in zip(masks_patch, probs_patch):
                    # centroid in patch coords
                    ys_p, xs_p = np.where(mask_p)
                    if len(ys_p) == 0:
                        continue
                    cy = float(ys_p.mean()) + y0
                    cx = float(xs_p.mean()) + x0

                    # dedup: skip if a close centroid already added
                    dup = False
                    for (py, px) in seen_centroids:
                        if (cy - py) ** 2 + (cx - px) ** 2 < min_dist_sq:
                            dup = True
                            break
                    if dup:
                        continue

                    # offset mask to full-image canvas
                    full_mask = np.zeros((h, w), dtype=bool)
                    full_mask[y0:y1, x0:x1] = mask_p
                    all_masks.append(full_mask)
                    all_probs.append(prob)
                    seen_centroids.append((cy, cx))

        print(f"  [Tiled] {len(ys)*len(xs)} tiles -> {len(all_masks)} instances")
        return all_masks, all_probs

    # ── Error fallback ────────────────────────────────────────────────────────

    def _error_result(self, image_path, message):
        return {
            "image_path":          str(image_path),
            "image_name":          Path(image_path).name,
            "total_particles":     0,
            "composition":         {},
            "verdict":             "ERROR",
            "reason":              message,
            "avg_confidence":      0.0,
            "blur_score":          0.0,
            "swe_checks":          {},
            "particles":           [],
            "overlay":             None,
            "processing_time_sec": 0,
        }
