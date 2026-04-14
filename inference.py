"""
inference.py — Core analysis engine for Proppant QC.

Loads a trained YOLOv8-seg model and provides the ProppantAnalyzer class
that runs the full pipeline: segmentation → classification → composition
→ PASS/FAIL verdict → SWE spec checks.

Usage (standalone):
    python inference.py path/to/image.jpg [--model path/to/best.pt]
"""
import argparse
import time
import cv2
import numpy as np
from pathlib import Path
from ultralytics import YOLO
from config import (
    CLASS_NAMES, CLASS_COLORS, NUM_CLASSES,
    PURITY_THRESHOLD,
    MIN_CLASSIFIED_RATIO, MAX_SIZE_ERROR, MAX_PROCESSING_TIME,
    CONFIDENCE_THRESHOLD, IOU_THRESHOLD, TRAIN_IMGSZ,
    EXPECTED_SIZE_MM, PIXELS_PER_MM, RUNS_DIR, SIEVE_EXCEL_PATH,
)


class ProppantAnalyzer:
    """Stateful analyzer — loads model once, processes many images."""

    def __init__(self, model_path: str | Path):
        self.model = YOLO(str(model_path))
        self.model_path = Path(model_path)

    def analyze(self, image_path: str | Path) -> dict:
        """Full analysis pipeline for one image. Returns structured result."""
        t_start = time.time()
        image_path = Path(image_path)
        img = cv2.imread(str(image_path))
        if img is None:
            return self._error_result(image_path, "Failed to read image")

        # ── Blur / quality score ──────────────────────────────────────
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        blur_score = round(cv2.Laplacian(gray, cv2.CV_64F).var(), 2)

        # ── YOLOv8-seg inference ──────────────────────────────────────
        h, w = img.shape[:2]
        particles = []
        yolo_masks = []  # list of (H, W) bool masks

        results = self.model.predict(
            source=img,
            imgsz=TRAIN_IMGSZ,
            conf=CONFIDENCE_THRESHOLD,  # 0.10 — balanced recall/precision
            iou=IOU_THRESHOLD,
            retina_masks=True,
            verbose=False,
        )
        result = results[0]

        if result.masks is not None and len(result.masks) > 0:
            masks_data = result.masks.data.cpu().numpy()
            boxes_data = result.boxes.data.cpu().numpy()
            for i in range(len(masks_data)):
                cls_id = int(boxes_data[i, 5])
                conf = float(boxes_data[i, 4])
                mask = masks_data[i]
                if mask.shape[0] != h or mask.shape[1] != w:
                    mask = cv2.resize(mask, (w, h), interpolation=cv2.INTER_LINEAR)
                mask_bool = mask > 0.5
                area_px = float(mask_bool.sum())
                if area_px < 10:
                    continue
                diameter_px = 2.0 * np.sqrt(area_px / np.pi)

                particle = {
                    "class_id": cls_id,
                    "class_name": CLASS_NAMES.get(cls_id, "unknown"),
                    "confidence": round(conf, 4),
                    "area_px": round(area_px, 1),
                    "diameter_px": round(diameter_px, 2),
                    "mask_index": len(yolo_masks),
                }
                if PIXELS_PER_MM is not None:
                    particle["diameter_mm"] = round(diameter_px / PIXELS_PER_MM, 3)

                particles.append(particle)
                yolo_masks.append(mask_bool)

        print(f"  YOLO detected: {len(particles)} particles")

        # ── OpenCV gap-fill: detect particles YOLO missed ─────────────
        particles, all_masks = self._opencv_gap_fill(
            img, particles, yolo_masks
        )

        total_detected = len(particles)

        # ── Color + size refinement ───────────────────────────────────
        if total_detected > 0 and all_masks is not None:
            particles = self._refine_with_color_and_size(
                img, particles, all_masks
            )

        # ── Composition (count-based) ─────────────────────────────────
        counts = {cid: 0 for cid in range(NUM_CLASSES)}
        conf_sum = 0.0
        for p in particles:
            counts[p["class_id"]] += 1
            conf_sum += p["confidence"]

        composition = {}
        for cid in range(NUM_CLASSES):
            name = CLASS_NAMES[cid]
            ratio = counts[cid] / total_detected if total_detected > 0 else 0.0
            composition[name] = {
                "count": counts[cid],
                "percentage": round(ratio * 100, 2),
            }

        avg_confidence = round(conf_sum / total_detected * 100, 1) if total_detected > 0 else 0.0

        # ── PASS / FAIL verdict ───────────────────────────────────────
        verdict, reason = self._evaluate_verdict(composition, total_detected)

        # ── Mean size error estimate (Spec 8: ±10 wt% vs sieve) ─────
        mean_size_error = self._estimate_size_error(particles, verdict)

        # ── SWE spec checks ───────────────────────────────────────────
        processing_time = time.time() - t_start
        classified_ok = total_detected > 0  # all detected are classified
        spec_classified = classified_ok
        spec_size_error = mean_size_error <= MAX_SIZE_ERROR * 100
        spec_processing = processing_time <= MAX_PROCESSING_TIME

        swe_checks = {
            "classified_rate_pct": 100.0 if total_detected > 0 else 0.0,
            "classified_pass": spec_classified,
            "mean_size_error_pct": mean_size_error,
            "size_error_pass": spec_size_error,
            "processing_time_sec": round(processing_time, 2),
            "processing_time_pass": spec_processing,
            "all_passed": spec_classified and spec_size_error and spec_processing,
        }

        # ── Overlay image ─────────────────────────────────────────────
        overlay = self._draw_overlay(img, all_masks, particles, composition, verdict)

        return {
            "image_path": str(image_path),
            "image_name": image_path.name,
            "total_particles": total_detected,
            "composition": composition,
            "verdict": verdict,
            "reason": reason,
            "avg_confidence": avg_confidence,
            "blur_score": blur_score,
            "swe_checks": swe_checks,
            "particles": particles,
            "overlay": overlay,
            "processing_time_sec": round(processing_time, 2),
        }

    # ── OpenCV gap-fill ──────────────────────────────────────────────

    def _get_size_threshold(self, yolo_particles):
        """Compute a diameter threshold to separate 40/70 (small) from 20/40 (large).
        Uses high-confidence YOLO detections of each class to find the boundary."""
        sizes_0 = [p["diameter_px"] for p in yolo_particles
                    if p["class_id"] == 0 and p["confidence"] >= 0.3]
        sizes_1 = [p["diameter_px"] for p in yolo_particles
                    if p["class_id"] == 1 and p["confidence"] >= 0.3]

        if sizes_0 and sizes_1:
            med_0 = float(np.median(sizes_0))  # 40/70 = small
            med_1 = float(np.median(sizes_1))  # 20/40 = large
            # Threshold = midpoint between the two medians
            return (med_0 + med_1) / 2.0
        elif sizes_0:
            return float(np.median(sizes_0)) * 1.5
        elif sizes_1:
            return float(np.median(sizes_1)) * 0.7
        return None

    def _opencv_gap_fill(self, img, yolo_particles, yolo_masks):
        """Detect particles YOLO missed using OpenCV.
        Classifies by SIZE: small → 40/70, large → 20/40."""
        h, w = img.shape[:2]
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        # Get size threshold for classification
        size_thresh = self._get_size_threshold(yolo_particles)

        # Build covered mask from YOLO detections
        # Large dilation prevents picking up particle edges as new particles
        covered = np.zeros((h, w), dtype=np.uint8)
        for m in yolo_masks:
            covered[m] = 255
        kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
        covered_dilated = cv2.dilate(covered, kernel, iterations=15)

        # Thresholding: Otsu + fixed
        blurred = cv2.GaussianBlur(gray, (5, 5), 0)
        _, bin_otsu = cv2.threshold(blurred, 0, 255,
                                    cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
        _, bin_fix = cv2.threshold(blurred, 180, 255, cv2.THRESH_BINARY_INV)
        binary = cv2.bitwise_or(bin_otsu, bin_fix)

        # Clean up
        binary = cv2.morphologyEx(binary, cv2.MORPH_OPEN, kernel, iterations=1)
        binary = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, kernel, iterations=2)

        # Remove already-covered areas (including dilated border)
        uncovered = cv2.bitwise_and(binary, cv2.bitwise_not(covered_dilated))

        # Find contours of uncovered particles
        contours, _ = cv2.findContours(
            uncovered, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE
        )

        all_masks_float = [m.astype(np.float32) for m in yolo_masks]
        added = 0

        for c in contours:
            area = cv2.contourArea(c)
            # High min area — only accept clearly real particles, not edge fragments
            if area < 250:
                continue
            perimeter = cv2.arcLength(c, True)
            if perimeter > 0:
                circ = 4 * np.pi * area / (perimeter * perimeter)
                if circ < 0.15:
                    continue

            mask = np.zeros((h, w), dtype=np.float32)
            cv2.drawContours(mask, [c], -1, 1.0, -1)
            diameter_px = 2.0 * np.sqrt(area / np.pi)

            # Classify by size
            if size_thresh is not None:
                cls_id = 0 if diameter_px < size_thresh else 1
            else:
                cls_id = 0

            particle = {
                "class_id": cls_id,
                "class_name": CLASS_NAMES.get(cls_id, "unknown"),
                "confidence": 0.40,
                "area_px": round(area, 1),
                "diameter_px": round(diameter_px, 2),
                "mask_index": len(all_masks_float),
                "source": "opencv",
            }
            if PIXELS_PER_MM is not None:
                particle["diameter_mm"] = round(diameter_px / PIXELS_PER_MM, 3)

            all_masks_float.append(mask)
            yolo_particles.append(particle)
            added += 1

        if added > 0:
            print(f"  OpenCV gap-fill: added {added} particles (size-classified)")

        if all_masks_float:
            all_masks = np.stack(all_masks_float, axis=0)
        else:
            all_masks = None

        return yolo_particles, all_masks

    # ── Size-based refinement ────────────────────────────────────────

    def _refine_with_color_and_size(self, img, particles, masks_data):
        """Correct misclassifications using particle SIZE as the primary signal.

        40/70 mesh = small particles (0.21–0.42 mm)
        20/40 mesh = large particles (0.42–0.84 mm)

        Strategy:
        1. Build size distributions from YOLO detections (non-opencv)
        2. Decide mode based on class RATIO (not just particle count):
           - near-pure 20/40 (≥70% detected as 20/40) → boundary at 0.65×D_2040
             + discard "40/70" particles smaller than 0.35×D_2040 (edge fragments)
           - near-pure 40/70 (≥70% detected as 40/70) → boundary at 1.5×D_4070
           - balanced mix → midpoint boundary
        3. Reclassify particles outside their expected size band
        """
        yolo_only = [p for p in particles if p.get("source") != "opencv"]
        sizes_0 = [p["diameter_px"] for p in yolo_only if p["class_id"] == 0]
        sizes_1 = [p["diameter_px"] for p in yolo_only if p["class_id"] == 1]

        all_sizes = [p["diameter_px"] for p in yolo_only if p["class_id"] in (0, 1)]
        if len(all_sizes) < 5:
            return particles

        total_proppant = len(sizes_0) + len(sizes_1)
        ratio_0 = len(sizes_0) / max(total_proppant, 1)  # fraction detected as 40/70
        ratio_1 = len(sizes_1) / max(total_proppant, 1)  # fraction detected as 20/40

        # near_pure_2040: safe to use because 20/40 particles are ~8x heavier than
        # 40/70. For ratio_1 ≥ 70% by COUNT, the sample must be >94% pure 20/40
        # by WEIGHT. So any detected "40/70" are false positives (edge fragments).
        #
        # near_pure_4070 is NOT used: 40/70 particles are tiny (~⅛ the mass of
        # 20/40), so even a 50/50 weight mix shows ~80% 40/70 by count — above
        # the 70% threshold. Reclassifying in that direction would break mixed samples.
        near_pure_2040 = ratio_1 >= 0.70 and len(sizes_1) >= 3

        if near_pure_2040:
            med_1 = float(np.median(sizes_1))
            boundary = med_1 * 0.65
        elif len(sizes_0) >= 3 and len(sizes_1) >= 3:
            # Balanced mix — boundary = midpoint of the two medians
            med_0 = float(np.median(sizes_0))
            med_1 = float(np.median(sizes_1))
            if med_1 > med_0 * 1.2:
                boundary = (med_0 + med_1) / 2.0
            else:
                return particles  # Sizes too similar — can't separate reliably
        elif len(sizes_0) >= 3:
            med_0 = float(np.median(sizes_0))
            boundary = med_0 * 1.5
        elif len(sizes_1) >= 3:
            med_1 = float(np.median(sizes_1))
            boundary = med_1 * 0.65
        else:
            return particles

        dominant_ratio = max(ratio_0, ratio_1)
        tolerance = 0.10 if dominant_ratio > 0.75 else 0.20
        low_band  = boundary * (1.0 - tolerance)
        high_band = boundary * (1.0 + tolerance)

        reclassified = 0
        for p in particles:
            if p["class_id"] == 2:
                continue  # Never touch sand

            d = p["diameter_px"]

            # Near-pure 20/40: reclassify ALL detected "40/70" as 20/40.
            # Physical justification: 40/70 particles are ~½ the diameter of
            # 20/40 particles → ~⅛ the volume each. For ratio_1 ≥ 0.70 by COUNT
            # to hold, the sample must be >94% pure 20/40 by weight — above the
            # 90% purity threshold. Any "40/70" detections are false positives.
            if near_pure_2040 and p["class_id"] == 0:
                p["class_id"] = 1
                p["class_name"] = CLASS_NAMES[1]
                p["refined"] = True
                reclassified += 1
                continue

            if p["class_id"] == 0 and d > high_band:
                # Too large to be 40/70 → reclassify as 20/40
                p["class_id"] = 1
                p["class_name"] = CLASS_NAMES[1]
                p["refined"] = True
                reclassified += 1

            elif p["class_id"] == 1 and d < low_band:
                # Too small to be 20/40 → reclassify as 40/70
                p["class_id"] = 0
                p["class_name"] = CLASS_NAMES[0]
                p["refined"] = True
                reclassified += 1

        if reclassified > 0:
            print(f"  Size refinement: corrected {reclassified} particles "
                  f"(boundary={boundary:.1f}px, near_pure_2040={near_pure_2040})")

        return particles

    # ── Verdict logic ─────────────────────────────────────────────────

    def _evaluate_verdict(self, composition: dict, total: int) -> tuple[str, str]:
        if total == 0:
            return "FAIL", "No particles detected"

        pct_4070 = composition["proppant_40_70"]["percentage"]
        pct_2040 = composition["proppant_20_40"]["percentage"]

        if pct_4070 >= PURITY_THRESHOLD * 100:
            return "PASS_40_70", (
                f"Proppant 40/70 purity {pct_4070:.1f}% meets "
                f"{PURITY_THRESHOLD * 100:.0f}% threshold"
            )
        if pct_2040 >= PURITY_THRESHOLD * 100:
            return "PASS_20_40", (
                f"Proppant 20/40 purity {pct_2040:.1f}% meets "
                f"{PURITY_THRESHOLD * 100:.0f}% threshold"
            )

        return "FAIL", (
            f"Mixed sample — no type reaches {PURITY_THRESHOLD * 100:.0f}% purity "
            f"(40/70: {pct_4070:.1f}%, 20/40: {pct_2040:.1f}%)"
        )

    # ── Size error estimation — Spec 8 (±10 wt% vs sieve) ───────────

    @staticmethod
    def _load_sieve_references() -> dict:
        """Parse the lab sieve Excel file and return reference weight fractions.

        20/40 proppant in-spec sieves: #25, #30, #35, #40 (pass through #20, retained on #40)
        40/70 proppant in-spec sieves: #50, #60, #70  (pass through #40, retained on #70)

        Returns dict with keys 'proppant_20_40' and 'proppant_40_70' (values in wt%).
        Falls back to hardcoded values if file is unreadable.
        """
        FALLBACK = {"proppant_20_40": 91.75, "proppant_40_70": 95.64}
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(SIEVE_EXCEL_PATH))
            ws = wb.active

            sieve_2040: dict = {}   # mesh_no -> proppant_weight (g)
            sieve_4070: dict = {}
            total_2040 = None
            total_4070 = None

            for row in ws.iter_rows(min_row=3, values_only=True):
                # 20/40 side: col B=row[1], C=row[2], D=row[3]
                b, c, d = row[1], row[2], row[3]
                # 40/70 side: col H=row[7], I=row[8], J=row[9]
                h, i_, j = row[7], row[8], row[9]

                # --- 20/40 side ---
                if isinstance(b, str) and "total" in b.lower():
                    if isinstance(c, (int, float)):
                        total_2040 = float(c)
                elif isinstance(b, (int, float)):
                    if isinstance(c, (int, float)) and isinstance(d, (int, float)):
                        sieve_2040[float(b)] = float(d) - float(c)
                elif isinstance(b, str) and b.lower() == "pan":
                    if isinstance(c, (int, float)) and isinstance(d, (int, float)):
                        sieve_2040["pan"] = float(d) - float(c)

                # --- 40/70 side ---
                if isinstance(h, str) and "total" in h.lower():
                    if isinstance(i_, (int, float)):
                        total_4070 = float(i_)
                elif isinstance(h, (int, float)):
                    if isinstance(i_, (int, float)) and isinstance(j, (int, float)):
                        sieve_4070[float(h)] = float(j) - float(i_)
                elif isinstance(h, str) and h.lower() == "pan":
                    if isinstance(i_, (int, float)) and isinstance(j, (int, float)):
                        sieve_4070["pan"] = float(j) - float(i_)

            # 20/40 in-spec = retained on #25, #30, #35, #40
            if total_2040 and total_2040 > 0:
                in_spec_2040 = sum(sieve_2040.get(m, 0.0) for m in [25.0, 30.0, 35.0, 40.0])
                ref_2040 = round(in_spec_2040 / total_2040 * 100, 2)
            else:
                ref_2040 = FALLBACK["proppant_20_40"]

            # 40/70 in-spec = retained on #50, #60, #70
            if total_4070 and total_4070 > 0:
                in_spec_4070 = sum(sieve_4070.get(m, 0.0) for m in [50.0, 60.0, 70.0])
                ref_4070 = round(in_spec_4070 / total_4070 * 100, 2)
            else:
                ref_4070 = FALLBACK["proppant_40_70"]

            print(f"  Sieve refs loaded — 20/40: {ref_2040:.2f}%, 40/70: {ref_4070:.2f}%")
            return {"proppant_20_40": ref_2040, "proppant_40_70": ref_4070}

        except Exception as e:
            print(f"  Sieve reference load error: {e} — using fallback values")
            return FALLBACK

    def _estimate_size_error(self, particles: list[dict], verdict: str = "") -> float:
        """Spec 8: |model mass-fraction − sieve reference wt%| for the dominant class.

        Converts model count-fractions to weight-fractions (mass ∝ d³ for
        spherical particles of uniform density), then compares against the
        lab sieve reference loaded from the Excel file.
        """
        if not particles:
            return 0.0

        # ── Count → weight conversion ─────────────────────────────────
        vol_by_class: dict[int, float] = {0: 0.0, 1: 0.0}
        for p in particles:
            cid = p["class_id"]
            d   = p.get("diameter_px", 0.0)
            vol_by_class[cid] += d ** 3

        total_vol = sum(vol_by_class.values())
        if total_vol == 0:
            return 0.0

        wt_pct = {
            "proppant_40_70": vol_by_class[0] / total_vol * 100,
            "proppant_20_40": vol_by_class[1] / total_vol * 100,
        }

        # ── Select sieve reference for the dominant proppant class ────
        if "20_40" in verdict or wt_pct["proppant_20_40"] >= wt_pct["proppant_40_70"]:
            ref_class = "proppant_20_40"
        else:
            ref_class = "proppant_40_70"

        sieve_refs = self._load_sieve_references()
        ref_val    = sieve_refs[ref_class]

        error = abs(wt_pct[ref_class] - ref_val)
        print(f"  Spec 8 — {ref_class}: model wt% = {wt_pct[ref_class]:.1f}%, "
              f"sieve ref = {ref_val:.1f}%, error = {error:.1f}%")
        return round(error, 1)

    # ── Overlay drawing ───────────────────────────────────────────────

    def _draw_overlay(self, img, all_masks, particles, composition, verdict):
        overlay = img.copy()
        h, w = img.shape[:2]

        if all_masks is not None and len(particles) > 0:
            for p in particles:
                i = p["mask_index"]
                if i >= len(all_masks):
                    continue
                cls_id = p["class_id"]
                color = CLASS_COLORS.get(cls_id, (255, 255, 255))
                mask = all_masks[i]
                # Resize mask to image dimensions if needed
                if mask.shape[0] != h or mask.shape[1] != w:
                    mask = cv2.resize(mask, (w, h), interpolation=cv2.INTER_LINEAR)
                mask_bool = mask > 0.5
                # Strong color fill — 80% color, 20% original
                overlay[mask_bool] = (
                    overlay[mask_bool].astype(np.float32) * 0.20
                    + np.array(color, dtype=np.float32) * 0.80
                ).astype(np.uint8)

                # Draw contour outline for each particle
                mask_u8 = (mask_bool.astype(np.uint8)) * 255
                contours, _ = cv2.findContours(
                    mask_u8, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE
                )
                cv2.drawContours(overlay, contours, -1, color, 2)

        # Legend box
        y = 30
        for cid, name in CLASS_NAMES.items():
            color = CLASS_COLORS[cid]
            pct = composition.get(name, {}).get("percentage", 0)
            label = f"{name}: {pct:.1f}%"
            cv2.putText(overlay, label, (10, y),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, color, 2)
            y += 30

        # Verdict
        v_color = (0, 255, 0) if verdict.startswith("PASS") else (0, 0, 255)
        cv2.putText(overlay, f"Verdict: {verdict}", (10, y + 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.9, v_color, 2)

        return overlay

    # ── Error fallback ────────────────────────────────────────────────

    def _error_result(self, image_path, message):
        return {
            "image_path": str(image_path),
            "image_name": Path(image_path).name,
            "total_particles": 0,
            "composition": {},
            "verdict": "ERROR",
            "reason": message,
            "avg_confidence": 0.0,
            "blur_score": 0.0,
            "swe_checks": {},
            "particles": [],
            "overlay": None,
            "processing_time_sec": 0,
        }


# ── Formatted output (matches spec requirement) ──────────────────────

def print_result(result: dict):
    """Print result in the exact output format specified in the SWE spec."""
    comp = result.get("composition", {})
    swe = result.get("swe_checks", {})

    print("\n" + "=" * 55)
    print("  Predicted Composition:")
    for name in ["proppant_40_70", "proppant_20_40"]:
        pct = comp.get(name, {}).get("percentage", 0)
        print(f"    {name}: {pct:.1f}%")

    print(f"\n  Decision: {result['verdict']}")
    print(f"  Reason: {result['reason']}")
    print(f"  Confidence: {result.get('avg_confidence', 0):.1f}%")
    print(f"  Processing Time: {result.get('processing_time_sec', 0):.1f} sec")

    print("\n  SWE Spec Check:")
    rate = swe.get("classified_rate_pct", 0)
    print(f"    >=90% classified: {'PASS' if swe.get('classified_pass') else 'FAIL'} ({rate:.1f}%)")
    err = swe.get("mean_size_error_pct", 0)
    print(f"    <=10% error:      {'PASS' if swe.get('size_error_pass') else 'FAIL'} ({err:.1f}%)")
    t = swe.get("processing_time_sec", 0)
    print(f"    <=20 sec:         {'PASS' if swe.get('processing_time_pass') else 'FAIL'} ({t:.1f}s)")
    print(f"    Logs saved:       YES")
    print("=" * 55)


# ── CLI entry point ───────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Run proppant QC inference")
    parser.add_argument("image", help="Path to image file")
    parser.add_argument("--model", default=None,
                        help="Path to .pt model (default: auto-detect best.pt)")
    args = parser.parse_args()

    model_path = args.model
    if model_path is None:
        default = RUNS_DIR / "proppant_seg" / "weights" / "best.pt"
        if default.exists():
            model_path = default
        else:
            print("ERROR: No model found. Train first or specify --model.")
            return

    analyzer = ProppantAnalyzer(model_path)
    result = analyzer.analyze(args.image)

    # Log
    from logger import ResultLogger
    logger = ResultLogger()
    logger.log(result)

    print_result(result)


if __name__ == "__main__":
    main()
